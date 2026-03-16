# pip install openpyxl
# This command installs the openpyxl library.
# openpyxl allows Python programs to read, write, and modify Excel (.xlsx) files.
# load_workbook is used to open an existing Excel workbook (file).
from openpyxl import load_workbook 



def read_excel_by_rows(file_path, sheet_name): # This function reads data from an Excel sheet row by row.
    workbook = load_workbook(file_path)  # Opens the Excel file located at the given file path.
    sheet = workbook[sheet_name] # Selects the specific sheet inside the Excel workbook.
    data = [] # Creating an empty list to store the rows of data.
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)      # Each row is returned as a tuple and added to the list.
    return data
    
    
    # iter_rows is a method that returns the rows of the sheet as tuples
    # iter_rows() loops through rows in the sheet.
    # min_row=2 skips the first row (usually header columns like username, password, etc.).
    # values_only=True is used to get only the cell values without any formatting


    #iter_cols can be used to iterate through columns

    # Example of directly accessing a specific cell:
    # firstRowFirstColumnValue = sheet.cell(row=2, column=1).value
    # firstRowSecondColumnValue = sheet.cell(row=2, column=2).value
    # return firstRowFirstColumnValue, firstRowSecondColumnValue  

def read_excel_by_cols(file_path, sheet_name):     # This function reads Excel data column by column instead of row by row.
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_cols(min_col=1, values_only=True):
        data.append(row)    
    return data

def addData_excel(file_path, sheet_name, rowNo, columnNo, value):     # This function updates or inserts a value into a specific cell in the Excel sheet.
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.cell(row=rowNo, column=columnNo).value = value     # Updates the value of the specified cell using row and column numbers.
    workbook.save(file_path)     # Saves the workbook so that the changes are written back to the Excel file.

def append_to_excel(file_path, sheet_name, headers, rows):     # This function appends (adds) new data at the bottom of the Excel sheet.
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.append(headers)
    # Adds a header row to the sheet.
    # Example: ["username", "password", "age"]
    for row in rows:
        sheet.append(row) # Adds each row of data at the end of the sheet.
    workbook.save(file_path)