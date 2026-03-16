from openpyxl import load_workbook
from utils.handlingExcel import read_excel_by_rows, read_excel_by_cols, addData_excel, append_to_excel
# Importing functions from our Excel utility file to handle reading and writing Excel data.


def test_printExcelbyRows():     # This test reads data from an Excel sheet **row by row** and prints it.
    filePath="testData\\searchProduct.xlsx"
    sheetName="PhoneConfigs"
    excelData = read_excel_by_rows(filePath, sheetName)     # Calls our utility function to get all rows from the sheet, skipping the header row
    print("data from excel in PhoneConfigs tab by rows is ", excelData)

def test_printExcelbyCols():     # This test reads data from an Excel sheet **column by column** and prints it.
    filePath="testData\\searchProduct.xlsx"
    sheetName="PhoneConfigs"
    excelData = read_excel_by_cols(filePath, sheetName)     # Calls our utility function to get all columns from the sheet
    print("data from excel in PhoneConfigs tab by cols is ", excelData)


def test_updateExcelValues():     # This test updates a specific cell in an Excel sheet.
    workbook = load_workbook("testData\\searchProduct.xlsx")
    sheet = workbook["MixerDetails"]
    print("from mixerdetails tab, the cell value is ", sheet.cell(row=3, column=1).value)
    #updating the cell value in excel
    sheet.cell(row=3, column=1,value = "NewDrinkMixerModelX")
    workbook.save("testData\\searchProduct.xlsx")
    print("updated value is ", sheet.cell(row=3, column=1).value)

def test_appendDataToExcel():     # This test appends new rows of data to an existing Excel sheet.
    filePath="testData\\searchProduct.xlsx"
    sheetName="MixerDetails"
    headers = ["Model", "Power", "Voltage"]      # Header row to be added to the sheet
    rows = [
        ("MixerModelY", "500W", "220V"),
        ("MixerMoabccccccdelY", "100500W", "22440V")
        ]
        # New rows to append
    append_to_excel(filePath, sheetName, headers, rows)     # Calls our utility function to append the header and rows to the sheet
    print("data appended to excel successfully")